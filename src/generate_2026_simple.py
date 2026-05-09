# -*- coding: utf-8 -*-
"""
Standalone-Generator für Lohnsteuertabelle 2026
Erstellt Excel-Datei direkt ohne externe venv-Abhängigkeiten
"""

import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from tax_data_2026 import KFB_VALUES_2026, TAX_PARAMS_2026
from tax_engine import (
    calculate_church_tax as calculate_church_tax_generic,
    calculate_monthly_tax_for_year,
    grundtarif,
    vorsorgepauschale,
)
from tax_year_config import (
    TAX_YEAR,
    get_generation_title,
    get_output_filename,
    get_sheet_names,
)

KFB_VALUES = KFB_VALUES_2026

def _grundtarif_2026(zvE: float) -> float:
    """§ 32a Abs. 1 EStG 2026"""
    return grundtarif(zvE, TAX_PARAMS_2026)


def _vorsorgepauschale(annual_gross: float) -> float:
    """Vereinfachte Vorsorgepauschale (GKV-AN, West 2026)."""
    return vorsorgepauschale(annual_gross, TAX_PARAMS_2026)


def calculate_monthly_tax(monthly_income: float, tax_class: int) -> tuple:
    """Berechnet monatliche Lohnsteuer und SolZ"""
    monthly_tax, monthly_solidarity = calculate_monthly_tax_for_year(monthly_income, tax_class, TAX_PARAMS_2026)
    return round(monthly_tax, 2), round(monthly_solidarity, 2)


def generate_excel_table(year: int = TAX_YEAR):
    """Erstellt die Excel-Datei mit der Lohnsteuertabelle des gewählten Jahres."""
    
    wb = Workbook()
    ws_main = wb.active
    main_sheet_name, raw_sheet_name = get_sheet_names(year)
    ws_main.title = main_sheet_name
    
    # Header
    ws_main['A1'] = "Einkommen_EUR"
    ws_main['B1'] = "Steuerklasse"
    ws_main['C1'] = "Lohnsteuer"
    
    col_idx = 4
    for kfb in KFB_VALUES:
        kfb_label = str(int(kfb)) if float(kfb).is_integer() else str(kfb)
        ws_main.cell(row=1, column=col_idx).value = f"KFB_{kfb_label}_SolZ"
        ws_main.cell(row=1, column=col_idx + 1).value = f"KFB_{kfb_label}_KiSt_9%"
        col_idx += 2
    
    # Styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for cell in ws_main[1]:
        if cell.value:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
    
    # Daten generieren
    row = 2
    income = 1000
    
    while income <= 10000 + 0.01:
        for tax_class in range(1, 7):
            ws_main.cell(row=row, column=1).value = income
            ws_main.cell(row=row, column=2).value = tax_class
            
            # Lohnsteuer (gleich für alle KFB-Werte bei dieser Einkommensstufe)
            lohnsteuer, _ = calculate_monthly_tax(income, tax_class)
            ws_main.cell(row=row, column=3).value = lohnsteuer
            
            # KFB-Werte
            col_idx = 4
            for kfb in KFB_VALUES:
                # Mit KFB wird das Einkommen reduziert
                kfb_reduction = kfb * 50  # 1 KFB ≈ 50€/Monat (vereinfacht)
                adjusted_income = max(0, income - kfb_reduction)
                
                _, solz = calculate_monthly_tax(adjusted_income, tax_class)
                kist = round(calculate_church_tax_generic(lohnsteuer), 2)  # 9% Kirchensteuer
                
                ws_main.cell(row=row, column=col_idx).value = solz
                ws_main.cell(row=row, column=col_idx + 1).value = kist
                col_idx += 2
            
            row += 1
        
        income += 5
    
    # Spaltenbreiten
    ws_main.column_dimensions['A'].width = 15
    ws_main.column_dimensions['B'].width = 14
    ws_main.column_dimensions['C'].width = 14
    for col in range(4, col_idx):
        ws_main.column_dimensions[get_column_letter(col)].width = 14
    
    # Rohdatenblatt (vereinfacht)
    ws_raw = wb.create_sheet(raw_sheet_name)
    ws_raw['A1'] = "Seite"
    ws_raw['B1'] = "Einkommen_EUR"
    ws_raw['C1'] = "Steuerklasse"
    ws_raw['D1'] = "Kinderfreibetrag"
    ws_raw['E1'] = "Lohnsteuer"
    ws_raw['F1'] = "SolZ"
    ws_raw['G1'] = "Kirchensteuer_9%"
    
    row = 2
    page = 1
    income = 1000
    
    while income <= 10000 + 0.01:
        for tax_class in range(1, 7):
            for kfb in KFB_VALUES:
                kfb_reduction = kfb * 50
                adjusted_income = max(0, income - kfb_reduction)
                
                lohnsteuer, solz = calculate_monthly_tax(income, tax_class)
                _, solz_adjusted = calculate_monthly_tax(adjusted_income, tax_class)
                kist = round(calculate_church_tax_generic(lohnsteuer), 2)
                
                ws_raw.cell(row=row, column=1).value = page
                ws_raw.cell(row=row, column=2).value = income
                ws_raw.cell(row=row, column=3).value = tax_class
                ws_raw.cell(row=row, column=4).value = kfb
                ws_raw.cell(row=row, column=5).value = lohnsteuer
                ws_raw.cell(row=row, column=6).value = solz_adjusted
                ws_raw.cell(row=row, column=7).value = kist
                row += 1
        
        income += 5
    
    # Spaltenbreiten für Rohdaten
    for col in range(1, 8):
        ws_raw.column_dimensions[get_column_letter(col)].width = 14
    
    return wb


def main():
    print("=" * 70)
    print(get_generation_title(TAX_YEAR))
    print("=" * 70)
    print("Einkommensbereiche: 1000€ - 10000€")
    print("Schrittweite: 5€")
    print("Steuerklassen: 1-6")
    print("Kinderfreibeträge: 0, 0.5, 1, 1.5, 2, 2.5, 3, 3.5, 4")
    print("Kirchensteuersatz: 9%")
    print("Region: West (monatlich)")
    print()
    
    # Ausgabedatei im aktuellen Verzeichnis oder src-Verzeichnis
    try:
        output_file = Path(__file__).parent / get_output_filename(TAX_YEAR)
    except:
        output_file = Path(f"d:/OneDrive/Git-Projekte/Lohnsteuertabellen-Konverter/{get_output_filename(TAX_YEAR)}")
    
    print("Generiere Excel-Datei...")
    wb = generate_excel_table(year=TAX_YEAR)
    
    print(f"Schreibe: {output_file}")
    wb.save(output_file)
    
    print()
    print("=" * 70)
    print("✓ FERTIG")
    print("=" * 70)
    print(f"Ausgabedatei: {output_file.absolute()}")
    
    return 0


if __name__ == "__main__":
    sys.exit(main())
